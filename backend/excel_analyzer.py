import os
import json
from typing import Any, Dict, List
from dotenv import load_dotenv

load_dotenv()

from ai_orchestrator import create_chat_completion

SYSTEM_PROMPT = """You are an expert data analyst specializing in Excel spreadsheet analysis. 
Your role is to:
1. Analyze uploaded Excel data and provide insights
2. Answer natural language questions about spreadsheet data
3. Generate summaries, trends, and key findings
4. Identify anomalies and patterns in the data
5. Provide actionable recommendations based on the data

Always provide clear, structured responses with specific data points and percentages when available.
Focus on being concise but comprehensive."""


def analyze_excel_data(file_path: str, user_query: str = None) -> Dict[str, Any]:
    """
    Analyze an Excel file using the AI orchestrator chain.

    Args:
        file_path: Path to the uploaded Excel file
        user_query: Optional user question about the data

    Returns:
        Analysis result with insights and findings
    """
    try:
        # Try to read Excel file and extract data
        import openpyxl
        import pandas as pd
        
        # First, try to load with pandas for better data handling
        try:
            excel_data = pd.read_excel(file_path)
            data_summary = excel_data.describe(include='all').to_string()
            data_preview = excel_data.head(10).to_string()
        except Exception:
            # Fallback to openpyxl
            workbook = openpyxl.load_workbook(file_path)
            ws = workbook.active
            data_preview = []
            for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
                data_preview.append(str(row))
            data_preview = '\n'.join(data_preview)
            data_summary = f"Loaded {ws.max_row} rows and {ws.max_column} columns"
        
        # Build the analysis prompt
        if user_query:
            prompt = f"""
Please analyze this Excel spreadsheet data and answer the following question:

Question: {user_query}

Data Summary:
{data_summary}

Data Preview:
{data_preview}

Provide a detailed analysis with specific insights, numbers, and trends from the data.
"""
        else:
            prompt = f"""
Please provide a comprehensive analysis of this Excel spreadsheet data:

Data Summary:
{data_summary}

Data Preview:
{data_preview}

Include:
1. Overview of what the data contains
2. Key statistics and trends
3. Notable patterns or anomalies
4. Actionable insights
"""
        
        messages = [
            {'role': 'system', 'content': SYSTEM_PROMPT},
            {'role': 'user', 'content': prompt}
        ]
        
        response = create_chat_completion(messages, max_tokens=2000)
        analysis = response.choices[0].message.content
        
        return {
            'success': True,
            'analysis': analysis,
            'file_path': file_path,
            'model': response.model,
            'query': user_query,
        }
    
    except ImportError:
        return {
            'success': False,
            'error': 'pandas or openpyxl not installed. Run: pip install pandas openpyxl',
            'analysis': None,
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'analysis': None,
        }


def query_spreadsheet_data(file_path: str, natural_language_query: str) -> Dict[str, Any]:
    """
    Answer a natural language question about spreadsheet data.
    
    Args:
        file_path: Path to the uploaded Excel file
        natural_language_query: User's question
    
    Returns:
        Answer with insights
    """
    return analyze_excel_data(file_path, natural_language_query)


def generate_spreadsheet_summary(file_path: str) -> Dict[str, Any]:
    """Generate a summary of the spreadsheet contents."""
    return analyze_excel_data(file_path, None)
